#!/usr/bin/env python3
"""Build the export workbook from a raw JSONL directory.

Usage:
    uv run --with openpyxl scripts/build_workbook.py <raw-dir> [--out <path>]

Reads candidates.jsonl / messages.jsonl / notes.jsonl (+ state.json for the
project slug) per contracts/raw-data-schemas.md and writes one workbook with
Candidates / Messages / Notes sheets per contracts/workbook-format.md.
Never overwrites an existing file: appends -HHMMSS to the name instead.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def fail(message: str) -> None:
    print(f"ERROR: {message}", file=sys.stderr)
    sys.exit(1)


def read_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        fail(f"missing input file: {path}")
    records = []
    with path.open(encoding="utf-8") as fh:
        for lineno, line in enumerate(fh, 1):
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as exc:
                fail(f"unparseable JSON at {path.name}:{lineno}: {exc}")
    return records


def dedupe_candidates(rows: list[dict]) -> list[dict]:
    by_url: dict[str, dict] = {}
    order: list[str] = []
    for row in rows:
        url = row.get("profile_url") or ""
        name = row.get("full_name") or ""
        if not url or not name:
            fail(f"invalid candidate record (empty profile_url or full_name): {row}")
        if url not in by_url:
            order.append(url)
        by_url[url] = row  # last occurrence wins
    return [by_url[url] for url in order]


def dedupe_by_key(rows: list[dict], key_fn) -> list[dict]:
    by_key: dict[tuple, dict] = {}
    order: list[tuple] = []
    for row in rows:
        key = key_fn(row)
        if key not in by_key:
            order.append(key)
        by_key[key] = row
    return [by_key[key] for key in order]


def check_attribution(rows: list[dict], known_urls: set[str], label: str) -> None:
    for row in rows:
        url = row.get("profile_url")
        if url not in known_urls:
            fail(f"{label} record references unknown profile_url: {url}")


def style_sheet(ws, widths: dict[int, int], wrap_cols: set[int]) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    if ws.max_row > 1:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column in wrap_cols:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(vertical="top")


def hyperlink(cell, url: str) -> None:
    if url:
        cell.hyperlink = url
        cell.style = "Hyperlink"


def resolve_output(out: Path) -> Path:
    if not out.exists():
        return out
    suffixed = out.with_name(f"{out.stem}-{datetime.now():%H%M%S}{out.suffix}")
    if suffixed.exists():
        fail(f"both {out.name} and {suffixed.name} already exist; wait a second and retry")
    return suffixed


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("raw_dir", type=Path)
    parser.add_argument("--out", type=Path, default=None)
    args = parser.parse_args()

    raw_dir: Path = args.raw_dir
    if not raw_dir.is_dir():
        fail(f"raw directory not found: {raw_dir}")

    state_path = raw_dir / "state.json"
    slug = raw_dir.name
    if state_path.exists():
        try:
            state = json.loads(state_path.read_text(encoding="utf-8"))
            slug = state.get("project", {}).get("slug") or slug
        except json.JSONDecodeError as exc:
            fail(f"unparseable state.json: {exc}")

    candidates = dedupe_candidates(read_jsonl(raw_dir / "candidates.jsonl"))
    messages = dedupe_by_key(
        read_jsonl(raw_dir / "messages.jsonl"),
        lambda r: (r.get("profile_url"), r.get("seq")),
    )
    notes = dedupe_by_key(
        read_jsonl(raw_dir / "notes.jsonl"),
        lambda r: (r.get("profile_url"), r.get("kind"), r.get("text")),
    )

    known_urls = {c["profile_url"] for c in candidates}
    check_attribution(messages, known_urls, "messages.jsonl")
    check_attribution(notes, known_urls, "notes.jsonl")

    msg_counts: dict[str, int] = {}
    for m in messages:
        msg_counts[m["profile_url"]] = msg_counts.get(m["profile_url"], 0) + 1
    note_counts: dict[str, int] = {}
    for n in notes:
        note_counts[n["profile_url"]] = note_counts.get(n["profile_url"], 0) + 1

    wb = Workbook()

    ws = wb.active
    ws.title = "Candidates"
    ws.append(["Name", "Profile URL", "Headline", "Company", "Location",
               "Stage", "Date Added", "Messages", "Notes/Tags"])
    for c in candidates:
        ws.append([c.get("full_name"), c.get("profile_url"), c.get("headline"),
                   c.get("company"), c.get("location"), c.get("stage"),
                   c.get("date_added"), msg_counts.get(c["profile_url"], 0),
                   note_counts.get(c["profile_url"], 0)])
        hyperlink(ws.cell(row=ws.max_row, column=2), c.get("profile_url"))
    style_sheet(ws, {1: 24, 2: 40, 3: 32, 4: 22, 5: 22, 6: 14, 7: 12, 8: 10, 9: 10},
                wrap_cols={3})

    ws = wb.create_sheet("Messages")
    ws.append(["Candidate", "Profile URL", "#", "Date", "Direction", "Subject", "Message"])
    for m in messages:
        ws.append([m.get("candidate_name"), m.get("profile_url"), m.get("seq"),
                   m.get("date"), m.get("direction"), m.get("subject"), m.get("text")])
        hyperlink(ws.cell(row=ws.max_row, column=2), m.get("profile_url"))
    style_sheet(ws, {1: 24, 2: 40, 3: 5, 4: 12, 5: 11, 6: 28, 7: 80}, wrap_cols={7})

    ws = wb.create_sheet("Notes")
    ws.append(["Candidate", "Profile URL", "Type", "Text", "Author", "Date"])
    for n in notes:
        ws.append([n.get("candidate_name"), n.get("profile_url"), n.get("kind"),
                   n.get("text"), n.get("author"), n.get("date")])
        hyperlink(ws.cell(row=ws.max_row, column=2), n.get("profile_url"))
    style_sheet(ws, {1: 24, 2: 40, 3: 8, 4: 70, 5: 16, 6: 12}, wrap_cols={4})

    default_out = Path("exports") / f"{slug}-{datetime.now():%Y-%m-%d}.xlsx"
    out = resolve_output(args.out if args.out else default_out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f"{out.resolve()}")
    print(f"Candidates: {len(candidates)}  Messages: {len(messages)}  Notes: {len(notes)}")


if __name__ == "__main__":
    main()
