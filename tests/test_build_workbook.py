"""Offline tests for scripts/build_workbook.py against tests/fixtures/sample-project."""

import json
import shutil
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
SCRIPT = REPO / ".claude/skills/linkedin-recruiter-export/scripts/build_workbook.py"
FIXTURE = REPO / "tests/fixtures/sample-project"


def run_builder(raw_dir: Path, out: Path):
    return subprocess.run(
        [sys.executable, str(SCRIPT), str(raw_dir), "--out", str(out)],
        capture_output=True, text=True,
    )


@pytest.fixture()
def workbook(tmp_path):
    out = tmp_path / "sample.xlsx"
    result = run_builder(FIXTURE, out)
    assert result.returncode == 0, result.stderr
    return load_workbook(out)


def test_sheets_and_columns(workbook):
    assert workbook.sheetnames == ["Candidates", "Messages", "Notes"]
    headers = [c.value for c in workbook["Candidates"][1]]
    assert headers == ["Name", "Profile URL", "Headline", "Company", "Location",
                       "Stage", "Date Added", "Messages", "Notes/Tags"]
    assert [c.value for c in workbook["Messages"][1]] == [
        "Candidate", "Profile URL", "#", "Date", "Direction", "Subject", "Message"]
    assert [c.value for c in workbook["Notes"][1]] == [
        "Candidate", "Profile URL", "Type", "Text", "Author", "Date"]
    assert workbook["Candidates"].freeze_panes == "A2"


def test_dedupe_keeps_last_occurrence(workbook):
    ws = workbook["Candidates"]
    rows = {ws.cell(r, 2).value: ws.cell(r, 3).value for r in range(2, ws.max_row + 1)}
    # fixture has sample-two twice; the later line has the updated headline
    assert len(rows) == 4
    assert rows["https://www.linkedin.com/in/sample-two"] == "مدير مالي أول"


def test_zero_message_candidate_absent_from_messages(workbook):
    ws = workbook["Messages"]
    urls = {ws.cell(r, 2).value for r in range(2, ws.max_row + 1)}
    assert "https://www.linkedin.com/in/sample-three" not in urls
    assert "https://www.linkedin.com/in/sample-four" not in urls


def test_message_counts_on_candidates_sheet(workbook):
    ws = workbook["Candidates"]
    counts = {ws.cell(r, 2).value: (ws.cell(r, 8).value, ws.cell(r, 9).value)
              for r in range(2, ws.max_row + 1)}
    assert counts["https://www.linkedin.com/in/sample-one"] == (3, 2)
    assert counts["https://www.linkedin.com/in/sample-three"] == (0, 0)


def test_arabic_roundtrip(workbook):
    ws = workbook["Candidates"]
    names = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
    assert "عبدالله السامي" in names
    msg_texts = [workbook["Messages"].cell(r, 7).value
                 for r in range(2, workbook["Messages"].max_row + 1)]
    assert any("فرصة مميزة" in (t or "") for t in msg_texts)


def test_scheduled_direction_preserved(workbook):
    ws = workbook["Messages"]
    directions = {ws.cell(r, 5).value for r in range(2, ws.max_row + 1)}
    assert "scheduled" in directions


def test_empty_project(tmp_path):
    raw = tmp_path / "empty-project"
    raw.mkdir()
    for name in ("candidates.jsonl", "messages.jsonl", "notes.jsonl"):
        (raw / name).write_text("", encoding="utf-8")
    out = tmp_path / "empty.xlsx"
    result = run_builder(raw, out)
    assert result.returncode == 0, result.stderr
    wb = load_workbook(out)
    for sheet in ("Candidates", "Messages", "Notes"):
        assert wb[sheet].max_row == 1  # headers only


def test_overwrite_protection(tmp_path):
    out = tmp_path / "clash.xlsx"
    first = run_builder(FIXTURE, out)
    assert first.returncode == 0
    second = run_builder(FIXTURE, out)
    assert second.returncode == 0
    produced = list(tmp_path.glob("clash*.xlsx"))
    assert len(produced) == 2  # suffixed file created, original untouched


def test_unknown_profile_url_fails(tmp_path):
    raw = tmp_path / "bad-project"
    shutil.copytree(FIXTURE, raw)
    with (raw / "messages.jsonl").open("a", encoding="utf-8") as fh:
        fh.write(json.dumps({"profile_url": "https://www.linkedin.com/in/ghost",
                             "candidate_name": "Ghost", "seq": 1, "date": None,
                             "direction": "sent", "subject": None, "text": "boo"}) + "\n")
    result = run_builder(raw, tmp_path / "bad.xlsx")
    assert result.returncode != 0
    assert "unknown profile_url" in result.stderr


def test_invalid_candidate_fails(tmp_path):
    raw = tmp_path / "invalid-project"
    shutil.copytree(FIXTURE, raw)
    with (raw / "candidates.jsonl").open("a", encoding="utf-8") as fh:
        fh.write(json.dumps({"profile_url": "", "full_name": "No URL"}) + "\n")
    result = run_builder(raw, tmp_path / "invalid.xlsx")
    assert result.returncode != 0
    assert "invalid candidate" in result.stderr
